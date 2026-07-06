from __future__ import annotations

import json
import math
import re
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output" / "spreadsheet"
OUT_PATH = OUT_DIR / "STS_Cairo_Potential_Leads_300.xlsx"
CACHE_DIR = ROOT / "tmp" / "lead_cache"
CACHE_PATH = CACHE_DIR / "cairo_osm_overpass.json"

OVERPASS_URLS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
    "https://overpass.openstreetmap.ru/api/interpreter",
]

# Greater Cairo east-bank / Cairo-focused bounding box. Includes New Cairo and
# main Cairo commercial districts while avoiding broad all-Egypt scraping.
BBOX = (29.82, 31.13, 30.20, 31.75)

INDUSTRY_RULES = [
    ("Medical Clinics", lambda t: t.get("amenity") in {"clinic", "doctors", "dentist"} or t.get("healthcare") in {"clinic", "doctor", "dentist"}),
    ("Hospitals & Healthcare", lambda t: t.get("amenity") == "hospital" or t.get("healthcare") == "hospital"),
    ("Pharmacies", lambda t: t.get("amenity") == "pharmacy" or t.get("shop") == "chemist"),
    ("Restaurants & Cafes", lambda t: t.get("amenity") in {"restaurant", "cafe", "fast_food", "food_court"}),
    ("Hotels & Hospitality", lambda t: t.get("tourism") in {"hotel", "guest_house", "hostel", "apartment"}),
    ("Real Estate & Property", lambda t: t.get("office") in {"estate_agent", "property_management"} or t.get("shop") == "estate_agent"),
    ("Education & Training", lambda t: t.get("amenity") in {"school", "college", "university", "kindergarten", "language_school"} or t.get("office") == "educational_institution"),
    ("Fitness & Wellness", lambda t: t.get("leisure") in {"fitness_centre", "sports_centre"} or t.get("amenity") in {"spa", "gym"}),
    ("Beauty & Personal Care", lambda t: t.get("shop") in {"beauty", "hairdresser", "cosmetics", "perfumery"}),
    ("Automotive", lambda t: t.get("shop") in {"car", "car_repair", "car_parts", "tyres"} or t.get("amenity") in {"fuel", "car_wash", "vehicle_inspection"}),
    ("Retail & E-commerce", lambda t: t.get("shop") in {"clothes", "shoes", "electronics", "mobile_phone", "computer", "department_store", "mall", "supermarket", "convenience", "gift", "furniture", "books", "jewelry", "sports"}),
    ("Professional Services", lambda t: t.get("office") in {"lawyer", "accountant", "company", "consulting", "insurance", "financial", "it", "telecommunication"}),
    ("Banks & Financial Services", lambda t: t.get("amenity") in {"bank", "bureau_de_change"} or t.get("office") == "financial"),
    ("Logistics & Industrial", lambda t: t.get("industrial") is not None or t.get("office") in {"logistics", "company"} or t.get("shop") == "trade"),
]

OFFER_BY_INDUSTRY = {
    "Medical Clinics": "Arabic WhatsApp booking assistant + website refresh + lead follow-up automation",
    "Hospitals & Healthcare": "Patient inquiry routing, appointment automation, and service landing pages",
    "Pharmacies": "WhatsApp reorder assistant, local SEO pages, and inventory request workflow",
    "Restaurants & Cafes": "Menu/booking website, WhatsApp ordering flow, review capture, and local ads funnel",
    "Hotels & Hospitality": "Booking inquiry assistant, multilingual landing pages, and guest follow-up automation",
    "Real Estate & Property": "Lead qualification bot, listing pages, CRM sheet, and automated follow-up",
    "Education & Training": "Admissions lead funnel, WhatsApp FAQ assistant, and parent/student follow-up",
    "Fitness & Wellness": "Membership lead funnel, booking automation, and retention reminders",
    "Beauty & Personal Care": "Booking website, Instagram-to-WhatsApp lead flow, and no-show reminders",
    "Automotive": "Service booking bot, quote workflow, and lead tracking dashboard",
    "Retail & E-commerce": "Product landing pages, WhatsApp sales assistant, and cart/review automation",
    "Professional Services": "Credibility website, consultation booking, and lead intake automation",
    "Banks & Financial Services": "Branch/service landing pages and customer inquiry routing",
    "Logistics & Industrial": "B2B website upgrade, quote intake workflow, and operations dashboard",
    "Other Local Business": "Website upgrade, WhatsApp lead capture, and simple CRM automation",
}

PRIORITY_WEIGHTS = {
    "website": 16,
    "phone": 14,
    "email": 14,
    "brand": 5,
    "opening_hours": 3,
}


def overpass_query() -> str:
    south, west, north, east = BBOX
    filters = [
        'node["name"]["amenity"]',
        'way["name"]["amenity"]',
        'relation["name"]["amenity"]',
        'node["name"]["shop"]',
        'way["name"]["shop"]',
        'relation["name"]["shop"]',
        'node["name"]["office"]',
        'way["name"]["office"]',
        'relation["name"]["office"]',
        'node["name"]["tourism"]',
        'way["name"]["tourism"]',
        'relation["name"]["tourism"]',
        'node["name"]["leisure"]',
        'way["name"]["leisure"]',
        'relation["name"]["leisure"]',
        'node["name"]["healthcare"]',
        'way["name"]["healthcare"]',
        'relation["name"]["healthcare"]',
        'node["name"]["industrial"]',
        'way["name"]["industrial"]',
        'relation["name"]["industrial"]',
    ]
    body = "\n".join(f"{f}({south},{west},{north},{east});" for f in filters)
    return f"""
[out:json][timeout:180];
(
{body}
);
out center tags;
"""


def fetch_data() -> dict[str, Any]:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    if CACHE_PATH.exists():
        age_hours = (time.time() - CACHE_PATH.stat().st_mtime) / 3600
        if age_hours < 24:
            return json.loads(CACHE_PATH.read_text(encoding="utf-8"))

    query = overpass_query()
    last_error = None
    for url in OVERPASS_URLS:
        try:
            response = requests.post(
                url,
                data={"data": query},
                headers={"User-Agent": "STS Cairo lead workbook generator"},
                timeout=220,
            )
            response.raise_for_status()
            data = response.json()
            CACHE_PATH.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
            return data
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            time.sleep(2)
    raise RuntimeError(f"Overpass fetch failed: {last_error}")


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def get_tag(tags: dict[str, Any], *names: str) -> str:
    for name in names:
        if tags.get(name):
            return clean(tags.get(name))
    return ""


def classify(tags: dict[str, Any]) -> str:
    for industry, rule in INDUSTRY_RULES:
        if rule(tags):
            return industry
    return "Other Local Business"


def lat_lon(element: dict[str, Any]) -> tuple[float | None, float | None]:
    if "lat" in element and "lon" in element:
        return element["lat"], element["lon"]
    center = element.get("center") or {}
    if "lat" in center and "lon" in center:
        return center["lat"], center["lon"]
    return None, None


def osm_url(element: dict[str, Any]) -> str:
    return f"https://www.openstreetmap.org/{element.get('type')}/{element.get('id')}"


def address(tags: dict[str, Any]) -> str:
    parts = [
        get_tag(tags, "addr:housenumber"),
        get_tag(tags, "addr:street"),
        get_tag(tags, "addr:suburb", "addr:district", "addr:neighbourhood"),
        get_tag(tags, "addr:city"),
    ]
    return ", ".join([p for p in parts if p])


def normalize_phone(phone: str) -> str:
    phone = clean(phone)
    phone = phone.replace(";", " / ")
    return phone


def lead_score(tags: dict[str, Any], industry: str) -> int:
    high_fit = {
        "Real Estate & Property",
        "Medical Clinics",
        "Restaurants & Cafes",
        "Hotels & Hospitality",
        "Education & Training",
        "Retail & E-commerce",
        "Professional Services",
        "Beauty & Personal Care",
        "Fitness & Wellness",
        "Automotive",
    }
    score = 58 if industry in high_fit else 48
    for tag, weight in PRIORITY_WEIGHTS.items():
        if tags.get(tag) or tags.get(f"contact:{tag}"):
            score += weight
    if tags.get("opening_hours"):
        score += 3
    if tags.get("name:en") and tags.get("name:ar"):
        score += 3
    return min(score, 100)


def priority(score: int) -> str:
    if score >= 82:
        return "A - Call first"
    if score >= 68:
        return "B - Strong fit"
    return "C - Research then call"


def first_offer(industry: str) -> str:
    return OFFER_BY_INDUSTRY.get(industry, OFFER_BY_INDUSTRY["Other Local Business"])


def business_type(tags: dict[str, Any]) -> str:
    for key in ["amenity", "shop", "office", "tourism", "leisure", "healthcare", "industrial"]:
        if tags.get(key):
            return f"{key}:{tags[key]}"
    return "mapped business"


def outreach_angle(industry: str, has_website: bool, has_phone: bool) -> str:
    if not has_website:
        return "Lead with a modern bilingual website + WhatsApp CTA audit."
    if not has_phone:
        return "Research direct phone, then pitch conversion tracking and lead capture."
    if industry in {"Real Estate & Property", "Medical Clinics", "Education & Training"}:
        return "Pitch WhatsApp qualification and automated follow-up for missed inquiries."
    if industry in {"Restaurants & Cafes", "Beauty & Personal Care", "Fitness & Wellness"}:
        return "Pitch booking/order automation and review capture."
    return "Pitch website upgrade plus a simple CRM and WhatsApp automation."


def rows_from_osm(data: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for element in data.get("elements", []):
        tags = element.get("tags") or {}
        name = get_tag(tags, "name:en", "name")
        if not name or len(name) < 2:
            continue

        industry = classify(tags)
        phone = normalize_phone(get_tag(tags, "contact:phone", "phone"))
        website = get_tag(tags, "contact:website", "website", "url")
        email = get_tag(tags, "contact:email", "email")
        lat, lon = lat_lon(element)
        addr = address(tags)
        if not addr and lat and lon:
            addr = "Cairo area coordinates available"

        key = re.sub(r"[^a-z0-9]+", "", f"{name}-{industry}-{lat}-{lon}".lower())
        if key in seen:
            continue
        seen.add(key)

        score = lead_score(tags, industry)
        rows.append(
            {
                "Company / Location": name,
                "Industry": industry,
                "Business Type": business_type(tags),
                "Priority": priority(score),
                "Lead Fit Score": score,
                "Recommended STS Offer": first_offer(industry),
                "Outreach Angle": outreach_angle(industry, bool(website), bool(phone)),
                "Phone": phone,
                "Email": email,
                "Website": website,
                "Address": addr,
                "Area / District": get_tag(tags, "addr:suburb", "addr:district", "addr:neighbourhood", "addr:city") or "Cairo",
                "Latitude": lat or "",
                "Longitude": lon or "",
                "Source": osm_url(element),
                "Data Notes": "Public OSM listing. Verify contact before outreach." if not phone and not email and not website else "Public OSM listing.",
            }
        )
    return rows


def balance_rows(rows: list[dict[str, Any]], target: int = 300) -> list[dict[str, Any]]:
    rows = sorted(rows, key=lambda r: (r["Lead Fit Score"], bool(r["Phone"]), bool(r["Website"])), reverse=True)
    by_industry: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_industry[row["Industry"]].append(row)

    chosen: list[dict[str, Any]] = []
    used_sources: set[str] = set()
    industries = sorted(by_industry, key=lambda k: len(by_industry[k]), reverse=True)

    # First pass: force variety.
    per_industry_cap = max(12, math.ceil(target / max(len(industries), 1)))
    for industry in industries:
        for row in by_industry[industry][:per_industry_cap]:
            if row["Source"] in used_sources:
                continue
            chosen.append(row)
            used_sources.add(row["Source"])
            if len(chosen) >= target:
                return chosen

    # Fill remaining with best-ranked leads.
    for row in rows:
        if row["Source"] in used_sources:
            continue
        chosen.append(row)
        used_sources.add(row["Source"])
        if len(chosen) >= target:
            break
    return chosen


def style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="0B1426")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2EC")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.auto_filter.ref = ws.dimensions


def autosize(ws, max_width: int = 55) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(value), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, length + 2)


def create_workbook(rows: list[dict[str, Any]], total_available: int) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cairo Leads"

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    style_sheet(ws)
    autosize(ws)
    ws.column_dimensions["F"].width = 46
    ws.column_dimensions["G"].width = 42
    ws.column_dimensions["K"].width = 36
    ws.column_dimensions["O"].width = 45
    ws.row_dimensions[1].height = 34

    table = Table(displayName="CairoLeads", ref=f"A1:{get_column_letter(len(headers))}{len(rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary.append(["Final lead count", len(rows)])
    summary.append(["Total public OSM records collected before filtering", total_available])
    summary.append(["Geography", "Cairo / Greater Cairo-focused bounding box"])
    summary.append(["Source", "OpenStreetMap via Overpass API"])
    summary.append(["Important note", "Contacts are public listing data where available; verify before outreach."])
    summary.append(["Best use", "Filter Priority A/B, then call or WhatsApp businesses with phone numbers first."])
    summary.append(["Generated date", time.strftime("%Y-%m-%d")])
    style_sheet(summary)
    autosize(summary, max_width=80)
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 90

    industry = wb.create_sheet("Industry Breakdown")
    counts = Counter(r["Industry"] for r in rows)
    industry.append(["Industry", "Lead Count"])
    for name, count in counts.most_common():
        industry.append([name, count])
    style_sheet(industry)
    autosize(industry)

    call_plan = wb.create_sheet("Call Plan")
    call_plan.append(["Priority", "Recommended first action", "Script angle"])
    call_plan.append(["A - Call first", "Call or WhatsApp same day", "You already have enough public info to personalize the pitch. Offer a quick website/lead-flow audit."])
    call_plan.append(["B - Strong fit", "Research decision maker, then call", "Reference their industry and pitch one practical automation."])
    call_plan.append(["C - Research then call", "Find phone/Instagram/Facebook first", "Use as a prospecting backlog; enrich before outreach."])
    style_sheet(call_plan)
    autosize(call_plan, max_width=90)

    wb.save(OUT_PATH)


def main() -> None:
    data = fetch_data()
    all_rows = rows_from_osm(data)
    selected = balance_rows(all_rows, 300)
    if len(selected) < 300:
        raise RuntimeError(f"Only found {len(selected)} usable leads; need at least 300.")
    create_workbook(selected, len(all_rows))
    print(OUT_PATH)
    print(f"selected={len(selected)} available={len(all_rows)}")


if __name__ == "__main__":
    main()
