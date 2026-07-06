from __future__ import annotations

import re
import shutil
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from create_cairo_leads_workbook import (
    fetch_data,
    get_tag,
    lat_lon,
    normalize_phone,
    osm_url,
)


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output" / "spreadsheet"
OUT_PATH = OUT_DIR / "STS_Cairo_SMB_Potential_Leads_300.xlsx"
DOWNLOAD_PATH = Path.home() / "Downloads" / OUT_PATH.name


CHAIN_RE = re.compile(
    r"\b("
    r"mcdonald|kfc|pizza hut|starbucks|costa|domino|subway|hardee|burger king|"
    r"buffalo burger|cook door|cilantro|beano|paul|carrefour|spinneys|metro market|"
    r"seoudi|kazyon|bim|hyper one|vodafone|orange|etisalat|telecom egypt|we store|"
    r"radio shack|adidas|nike|hm|h&m|zara|pull&bear|bershka|mango|lc waikiki|"
    r"cib|qnb|hsbc|aaib|emirates nbd|national bank|banque|credit agricole|"
    r"alex bank|faisal|arab bank|saib|adib|bmw|mercedes|toyota|nissan|honda|"
    r"hyundai|kia|suzuki|volkswagen|chevrolet|audi|shell|mobil|totalenergies|"
    r"wataneya|watanya|car gas|total|emarat|on the run"
    r")\b",
    re.IGNORECASE,
)

EXCLUDE_NAME_RE = re.compile(
    r"\b("
    r"embassy|ministry|authority|government|governorate|museum|mosque|church|"
    r"cathedral|monastery|hospital|university|faculty|bank|atm|library|"
    r"police|station|club house|mall|hypermarket"
    r")\b",
    re.IGNORECASE,
)

TARGET_CAPS = {
    "Restaurants & Cafes": 55,
    "Specialty Retail": 45,
    "Medical & Dental Clinics": 38,
    "Beauty & Personal Care": 34,
    "Fitness & Wellness": 30,
    "Professional Services": 30,
    "Automotive Services": 28,
    "Pharmacies": 24,
    "Real Estate Offices": 22,
    "Education & Training": 20,
    "Boutique Hospitality": 14,
}

OFFERS = {
    "Restaurants & Cafes": "Menu/booking website + WhatsApp ordering + review capture",
    "Specialty Retail": "Product landing pages + WhatsApp sales assistant + simple CRM",
    "Medical & Dental Clinics": "Arabic WhatsApp booking assistant + clinic website refresh",
    "Beauty & Personal Care": "Booking website + Instagram-to-WhatsApp lead flow + reminders",
    "Fitness & Wellness": "Membership lead funnel + WhatsApp booking + renewal reminders",
    "Professional Services": "Credibility website + consultation booking + intake automation",
    "Automotive Services": "Service booking workflow + quote intake + follow-up dashboard",
    "Pharmacies": "WhatsApp reorder assistant + local SEO pages + inventory request workflow",
    "Real Estate Offices": "Lead qualification bot + listing pages + CRM follow-up",
    "Education & Training": "Admissions lead funnel + WhatsApp FAQ + follow-up automation",
    "Boutique Hospitality": "Booking inquiry assistant + direct booking landing page",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def classify_smb(tags: dict[str, Any], name: str) -> str | None:
    amenity = clean(tags.get("amenity")).lower()
    shop = clean(tags.get("shop")).lower()
    office = clean(tags.get("office")).lower()
    tourism = clean(tags.get("tourism")).lower()
    leisure = clean(tags.get("leisure")).lower()
    healthcare = clean(tags.get("healthcare")).lower()
    name_l = name.lower()

    if amenity in {"restaurant", "cafe", "fast_food", "ice_cream"}:
        return "Restaurants & Cafes"
    if amenity in {"clinic", "doctors", "dentist"} or healthcare in {"clinic", "doctor", "dentist"}:
        return "Medical & Dental Clinics"
    if amenity == "pharmacy" or shop == "chemist":
        return "Pharmacies"
    if shop in {"hairdresser", "beauty", "cosmetics", "perfumery", "massage"}:
        return "Beauty & Personal Care"
    if leisure in {"fitness_centre", "sports_centre"} or amenity in {"gym", "spa"}:
        return "Fitness & Wellness"
    if shop in {"car_repair", "car_parts", "tyres"} or amenity in {"car_wash"}:
        return "Automotive Services"
    if office in {"lawyer", "accountant", "consulting", "company", "insurance", "it", "telecommunication", "architect", "employment_agency"}:
        return "Professional Services"
    if office in {"estate_agent", "property_management"} or shop == "estate_agent":
        return "Real Estate Offices"
    if shop in {
        "clothes",
        "shoes",
        "electronics",
        "mobile_phone",
        "computer",
        "furniture",
        "jewelry",
        "gift",
        "sports",
        "books",
        "florist",
        "optician",
        "stationery",
        "pet",
        "baby_goods",
        "interior_decoration",
        "kitchen",
        "appliance",
        "lighting",
    }:
        return "Specialty Retail"
    if amenity in {"language_school", "music_school", "driving_school"}:
        return "Education & Training"
    if amenity == "school" and re.search(r"academy|center|centre|training|institute|nursery|preschool|language", name_l):
        return "Education & Training"
    if tourism in {"guest_house", "hostel", "apartment"}:
        return "Boutique Hospitality"
    if tourism == "hotel" and not re.search(r"hilton|marriott|sheraton|fairmont|kempinski|ritz|steigenberger|four seasons|intercontinental|novotel|holiday inn", name_l):
        return "Boutique Hospitality"
    return None


def exclusion_reason(tags: dict[str, Any], name: str) -> str:
    amenity = clean(tags.get("amenity")).lower()
    shop = clean(tags.get("shop")).lower()
    office = clean(tags.get("office")).lower()
    tourism = clean(tags.get("tourism")).lower()

    if CHAIN_RE.search(name):
        return "known chain / enterprise brand"
    if EXCLUDE_NAME_RE.search(name):
        return "institutional or enterprise-sounding listing"
    if amenity in {"bank", "atm", "hospital", "university", "fuel", "place_of_worship", "library", "police"}:
        return f"excluded amenity:{amenity}"
    if shop in {"supermarket", "mall", "department_store", "convenience"}:
        return f"excluded shop:{shop}"
    if office in {"government", "association", "ngo"}:
        return f"excluded office:{office}"
    if tourism in {"museum", "attraction"}:
        return f"excluded tourism:{tourism}"
    return ""


def address(tags: dict[str, Any]) -> str:
    parts = [
        get_tag(tags, "addr:housenumber"),
        get_tag(tags, "addr:street"),
        get_tag(tags, "addr:suburb", "addr:district", "addr:neighbourhood"),
        get_tag(tags, "addr:city"),
    ]
    return ", ".join([p for p in parts if p])


def business_type(tags: dict[str, Any]) -> str:
    for key in ["amenity", "shop", "office", "tourism", "leisure", "healthcare"]:
        if tags.get(key):
            return f"{key}:{clean(tags[key])}"
    return "local business"


def score_lead(row: dict[str, Any]) -> int:
    score = 52
    if row["Phone"]:
        score += 18
    if row["Website"]:
        score += 14
    if row["Email"]:
        score += 10
    if row["Address"] and row["Address"] != "Cairo area coordinates available":
        score += 4
    if row["Industry"] in {"Medical & Dental Clinics", "Real Estate Offices", "Beauty & Personal Care", "Fitness & Wellness"}:
        score += 6
    if row["Industry"] in {"Restaurants & Cafes", "Specialty Retail", "Professional Services"}:
        score += 4
    return min(score, 100)


def priority(score: int) -> str:
    if score >= 82:
        return "A - Call first"
    if score >= 68:
        return "B - Strong fit"
    return "C - Enrich first"


def smb_reason(industry: str, row: dict[str, Any]) -> str:
    if row["Phone"] and row["Website"]:
        return f"Targeted {industry.lower()} listing with public phone and website."
    if row["Phone"]:
        return f"Targeted {industry.lower()} listing with public phone."
    if row["Website"]:
        return f"Targeted {industry.lower()} listing with public website."
    return f"Targeted {industry.lower()} listing; enrich phone/Instagram before outreach."


def outreach_angle(industry: str, has_phone: bool, has_site: bool) -> str:
    if industry in {"Medical & Dental Clinics", "Beauty & Personal Care", "Fitness & Wellness"}:
        return "Pitch missed-call/WhatsApp booking automation and appointment reminders."
    if industry == "Restaurants & Cafes":
        return "Pitch a menu/booking landing page, WhatsApp ordering, and review capture."
    if industry == "Real Estate Offices":
        return "Pitch lead qualification, listing pages, and automated follow-up."
    if industry == "Specialty Retail":
        return "Pitch product landing pages and a WhatsApp sales assistant."
    if industry == "Professional Services":
        return "Pitch credibility website, consultation booking, and intake automation."
    if not has_phone:
        return "Find phone or Instagram first, then offer a quick lead-flow audit."
    if not has_site:
        return "Lead with a website upgrade and WhatsApp CTA audit."
    return "Offer a 10-minute audit of website, WhatsApp response, and follow-up flow."


def collect_rows(data: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    excluded: list[dict[str, Any]] = []
    seen: set[str] = set()

    for element in data.get("elements", []):
        tags = element.get("tags") or {}
        name = get_tag(tags, "name:en", "name")
        if not name:
            continue

        reason = exclusion_reason(tags, name)
        if reason:
            excluded.append({"Name": name, "Reason": reason, "Source": osm_url(element)})
            continue

        industry = classify_smb(tags, name)
        if not industry:
            continue

        lat, lon = lat_lon(element)
        phone = normalize_phone(get_tag(tags, "contact:phone", "phone"))
        email = get_tag(tags, "contact:email", "email")
        website = get_tag(tags, "contact:website", "website", "url")
        addr = address(tags)
        if not addr and lat and lon:
            addr = "Cairo area coordinates available"

        key = re.sub(r"[^a-z0-9\u0600-\u06ff]+", "", f"{name}-{industry}-{round(lat or 0, 4)}-{round(lon or 0, 4)}".lower())
        if key in seen:
            continue
        seen.add(key)

        row = {
            "Company / Location": name,
            "Industry": industry,
            "Business Type": business_type(tags),
            "Priority": "",
            "Lead Fit Score": 0,
            "SMB Fit Reason": "",
            "Recommended STS Offer": OFFERS[industry],
            "Outreach Angle": "",
            "Phone": phone,
            "Email": email,
            "Website": website,
            "Address": addr,
            "Area / District": get_tag(tags, "addr:suburb", "addr:district", "addr:neighbourhood", "addr:city") or "Cairo",
            "Latitude": lat or "",
            "Longitude": lon or "",
            "Source": osm_url(element),
            "Data Notes": "SMB prospect inferred from public listing category/name. Verify owner/contact before outreach.",
        }
        row["Lead Fit Score"] = score_lead(row)
        row["Priority"] = priority(row["Lead Fit Score"])
        row["SMB Fit Reason"] = smb_reason(industry, row)
        row["Outreach Angle"] = outreach_angle(industry, bool(phone), bool(website))
        rows.append(row)

    return rows, excluded


def select_300(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = sorted(rows, key=lambda r: (r["Lead Fit Score"], bool(r["Phone"]), bool(r["Website"])), reverse=True)
    by_industry: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_industry[row["Industry"]].append(row)

    selected: list[dict[str, Any]] = []
    used: set[str] = set()
    for industry, cap in TARGET_CAPS.items():
        for row in by_industry.get(industry, [])[:cap]:
            selected.append(row)
            used.add(row["Source"])

    if len(selected) < 300:
        for row in rows:
            if row["Source"] in used:
                continue
            selected.append(row)
            used.add(row["Source"])
            if len(selected) >= 300:
                break

    return selected[:300]


def style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="0B1426")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2EC")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def autosize(ws, max_width: int = 58) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(value), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, length + 2)


def add_table(ws, name: str) -> None:
    table = Table(displayName=name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def create_workbook(selected: list[dict[str, Any]], candidates: list[dict[str, Any]], excluded: list[dict[str, Any]]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "SMB Leads"
    headers = list(selected[0].keys())
    ws.append(headers)
    for row in selected:
        ws.append([row[h] for h in headers])
    style_sheet(ws)
    add_table(ws, "SMBLeads")
    autosize(ws)
    for col in ["F", "G", "H", "L", "P", "Q"]:
        ws.column_dimensions[col].width = 44
    ws.row_dimensions[1].height = 34

    summary = wb.create_sheet("Summary")
    counts = Counter(row["Industry"] for row in selected)
    priority_counts = Counter(row["Priority"] for row in selected)
    summary_rows = [
        ("Metric", "Value"),
        ("Final SMB lead count", len(selected)),
        ("SMB candidates after strict filtering", len(candidates)),
        ("Public listings excluded as broad/enterprise/institutional", len(excluded)),
        ("Industries included", len(counts)),
        ("Phone numbers available", sum(1 for row in selected if row["Phone"])),
        ("Websites available", sum(1 for row in selected if row["Website"])),
        ("Emails available", sum(1 for row in selected if row["Email"])),
        ("A priority leads", priority_counts.get("A - Call first", 0)),
        ("B priority leads", priority_counts.get("B - Strong fit", 0)),
        ("C priority leads", priority_counts.get("C - Enrich first", 0)),
        ("Source", "OpenStreetMap public business/place listings via Overpass API"),
        ("Quality note", "This version excludes banks, hospitals, universities, fuel stations, supermarkets/malls, government/institutional listings, and obvious chains."),
        ("Generated date", time.strftime("%Y-%m-%d")),
    ]
    for row in summary_rows:
        summary.append(row)
    style_sheet(summary)
    autosize(summary, 95)
    summary.column_dimensions["B"].width = 100

    breakdown = wb.create_sheet("Industry Breakdown")
    breakdown.append(["Industry", "Lead Count", "With Phone", "With Website", "With Email"])
    for industry, count in counts.most_common():
        industry_rows = [row for row in selected if row["Industry"] == industry]
        breakdown.append([
            industry,
            count,
            sum(1 for row in industry_rows if row["Phone"]),
            sum(1 for row in industry_rows if row["Website"]),
            sum(1 for row in industry_rows if row["Email"]),
        ])
    style_sheet(breakdown)
    add_table(breakdown, "IndustryBreakdown")
    autosize(breakdown)

    call_plan = wb.create_sheet("Call Plan")
    call_plan.append(["Priority", "Use these first", "Recommended action"])
    call_plan.append(["A - Call first", "Phone/email/website usually available", "Call or WhatsApp with a 10-minute audit offer and one industry-specific automation idea."])
    call_plan.append(["B - Strong fit", "Good category fit but may need enrichment", "Find Instagram/Facebook/decision maker, then call with the same offer."])
    call_plan.append(["C - Enrich first", "Good SMB category but missing direct contact", "Use Google Maps/Facebook to enrich phone before outreach."])
    call_plan.append(["Suggested first pitch", "Website + WhatsApp lead capture", "Ask: 'Do you currently lose inquiries from WhatsApp/Instagram because no one follows up fast enough?'"])
    style_sheet(call_plan)
    autosize(call_plan, 95)

    excluded_ws = wb.create_sheet("Exclusion Logic")
    excluded_ws.append(["Excluded Example / Name", "Reason", "Source"])
    for row in excluded[:250]:
        excluded_ws.append([row["Name"], row["Reason"], row["Source"]])
    style_sheet(excluded_ws)
    add_table(excluded_ws, "ExclusionExamples")
    autosize(excluded_ws, 80)

    wb.save(OUT_PATH)
    DOWNLOAD_PATH.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(OUT_PATH, DOWNLOAD_PATH)


def main() -> None:
    data = fetch_data()
    candidates, excluded = collect_rows(data)
    selected = select_300(candidates)
    if len(selected) < 300:
        raise RuntimeError(f"Only found {len(selected)} strict SMB leads.")
    create_workbook(selected, candidates, excluded)
    print(OUT_PATH)
    print(DOWNLOAD_PATH)
    print(f"selected={len(selected)} candidates={len(candidates)} excluded={len(excluded)}")
    print(Counter(row["Industry"] for row in selected).most_common())


if __name__ == "__main__":
    main()
